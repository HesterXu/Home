# -*- coding: utf-8 -*-
# @Time     : 2018/12/13/9:58
# @Author   : Hester Xu
# Email     : 1603046502@qq.com
# @File     : do_excel.py
# @Software : PyCharm

"""操作excel"""
from openpyxl import Workbook,load_workbook
import string
# from Util.FormatTime import data_time_chinese

class Excel_r_w:
    def __init__(self,excel_path,sheet_name):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path)  # 打开工作簿
        self.ws = self.wb[sheet_name]        # 定位表单
        # 以下为获取多个列的列号组成的列表，不论列是否有数据
        uper_str = string.ascii_uppercase
        self.col_no_list = list(map(lambda x: x, uper_str)) # 获取基础列号列表，即：['A','B'...'Z']
        for iA in 'ABCD': # 在基础列号列表后面加上更多的列，即：['A'...'Z';'AA','AB'...'AZ';'BA','BB'...'BZ'...]
            for iu in uper_str:
                self.col_no_list.append(iA + iu)

    # 获取sheet中的最大行数：
    def get_max_row(self):
        return self.ws.max_row

    # 获取sheet中的最大列数：
    def get_max_col(self):
        sol_no = self.ws.max_colimn  # 先获得列号
        # print("filter=",list(filter(lambda x: x==self.col_no_list[sol_no-1],self.col_no_list)))
        return list(filter(lambda x: x==self.col_no_list[sol_no-1],self.col_no_list))[0] # 再转成字母

    # 获取sheet的最小（起始）行号：
    def get_min_row(self):
        return self.ws.min_row

    # 获取sheet的最小（起始）列号：
    def get_min_col(self):
        sol_no = self.ws.min_column   # 先获得列号
        return list(filter(lambda x: x == self.col_no_list[sol_no - 1], self.col_no_list))[0]  # 再转成字母





