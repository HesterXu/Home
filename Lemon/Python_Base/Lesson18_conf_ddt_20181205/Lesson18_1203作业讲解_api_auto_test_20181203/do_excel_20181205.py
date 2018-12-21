# -*- coding: utf-8 -*-
# @Time     : 2018/12/6/6:26
# @Author   : Hester Xu
# Email     : xuruizhu@yeah.net
# @File     : do_excel_20181205.py
# @Software : PyCharm

from openpyxl import load_workbook
import pprint

class DoExcel:
    def get_data(self):
        wb = load_workbook('test_cases.xlsx')
        sheet = wb['test_data']

        test_data = []
        for i in range(2,sheet.max_row+1):
            row_data = {}
            row_data['case_id'] = sheet.cell(i,1).value
            row_data['title'] = sheet.cell(i,2).value
            row_data['url'] = sheet.cell(i,3).value
            row_data['param'] = sheet.cell(i,4).value
            row_data['http_method'] = sheet.cell(i,5).value
            row_data['expected'] = sheet.cell(i,6).value
            test_data.append(row_data)
        return test_data

#测试代码
if __name__ == '__main__':
    test_data = DoExcel().get_data()
    pprint.pprint(test_data)
