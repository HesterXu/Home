# -*- coding: utf-8 -*-
# @Time     : 2018/12/5/9:52
# @Author   : Hester Xu
# Email     : xuruizhu@yeah.net
# @File     : Lesson17_do_excel_02_20181203.py
# @Software : PyCharm

"""如何把每一行的数据读取到一个字典中，所有行的数据以字典的格式存在一个列表中"""
from openpyxl import load_workbook
wb = load_workbook('python12.xlsx')  # 打开excel
sheet = wb['Sheet1']  # 定位表单
test_data = []        # 定义一个空列表         # --复盘！！！
for i in range(1,sheet.max_row+1):
    sub_data = {}     # 定义一个空字典         # --复盘！！！
    sub_data['A'] = sheet.cell(i,1).value
    sub_data['B'] = sheet.cell(i,2).value
    sub_data['C'] = sheet.cell(i,3).value
    sub_data['D'] = sheet.cell(i,4).value
    sub_data['E'] = sheet.cell(i,5).value
    sub_data['F'] = sheet.cell(i,6).value
    sub_data['G'] = sheet.cell(i,7).value
    sub_data['H'] = sheet.cell(i,8).value
    sub_data['I'] = sheet.cell(i,9).value
    test_data.append(sub_data)   # 把sub_data添加到列表中    # --复盘！！！
""" test_data列表，是由一个个字典sub_data组成。
有几行，就有几个字典。每个字典中，key是列的值A,B,C等，value是该单元格的数据。"""
print(test_data)



















