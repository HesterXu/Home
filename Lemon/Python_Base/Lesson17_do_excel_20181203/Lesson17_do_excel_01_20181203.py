# -*- coding: utf-8 -*-
# @Time     : 2018/12/5/10:30
# @Author   : Hester Xu
# Email     : xuruizhu@yeah.net
# @File     : Lesson17_do_excel_01_20181203.py
# @Software : PyCharm

from openpyxl import load_workbook  #可以对excel进行写操作

# 1.打开excel      把要打开的excel表存到work_book工作簿中
work_book = load_workbook('python12.xlsx')  # 直接打开，相对路径  返回工作簿  #文件路径--要去复盘！！！

# 2.定位表单       把要打开的sheet存到sheet表单中
sheet = work_book['Sheet1']  #和字典类似      返回表单
# sheet_1 = work_book.get_sheet_by_name('sheet1')  # 已过时，会报警告信息

# 3.读取大佬名字  把要读取的名字存到dalao中
dalao_1 = sheet.cell(1,2).value   # 读数据行列值，从1开始
dalao_2 = sheet.cell(2,3).value
dalao_3 = sheet.cell(3,4).value
dalao_4 = sheet.cell(4,5).value
dalao_5 = sheet.cell(5,6).value
dalao_6 = sheet.cell(2,6).value

# 注意：excel中的数据类型变化
"""整数/浮点数/字符串：类型不变，字典/元组/列表：类型变为字符串"""
print(dalao_1,type(dalao_1))  # 整数--》整数
print(dalao_2,type(dalao_2))  # 字典--》字符串
print(dalao_3,type(dalao_3))  # 浮点数--》浮点数
print(dalao_4,type(dalao_4))  # 列表--》字符串
print(dalao_5,type(dalao_5))  # 元组--》字符串
print(dalao_6,type(dalao_6))  # 字符串--》字符串

# 4.怎么获取最大的行列值？
max_row = sheet.max_row      #最大行
max_col = sheet.max_column   #最大列
print('行：',max_row)
print('列：',max_col)

# 写数据
sheet.cell(4,4).value = 'Nicky' # 赋值，替换原来的值，达到写数据
work_book.save('python12.xlsx')  # 保存工作簿

work_book.save('python14.xlsx')  # 另存

# 增加数据
sheet.cell(6,9).value = 'Andy' # 增加一个单元格的数据
work_book.save('python12.xlsx')  # 保存工作簿

# 怎么新建excel
from openpyxl import Workbook

wb = Workbook('python13.xlsx')    # 新建一个excel
wb.save('python13.xlsx')          # 保存工作簿

file = open('pythontest.xlsx','w+')  # 新建一个excel，打不开？？？
