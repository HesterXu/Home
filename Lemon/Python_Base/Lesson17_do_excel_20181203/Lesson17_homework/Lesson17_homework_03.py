# -*- coding: utf-8 -*-
# @Time     : 2018/12/5/11:02
# @Author   : Hester Xu
# Email     : xuruizhu@yeah.net
# @File     : Lesson17_homework_03.py
# @Software : PyCharm

"""1203-在1130作业的基础上完成测试用例的参数化
1：根据老师上课讲解的测试用例参数化，去写单元测试用例以及用到超继承
2：把所有的测试用例数据存到Excel中，然后写一个类do_excel，完成测试数据的读取
3：利用do_excel类读取到的测试数据 传给1203课堂上讲解的test_suite，完成测试用例的加载。
4：要求：登录3条用例 充值3条用例
5：请给用例加上断言 以及异常处理，断言可以用code 也可以用msg或者是用整段信息都OK，最后要有测试报告出具
"""

import unittest
import HTMLTestRunnerNew
from Lemon.Python_Base.Lesson17_do_excel_20181203.Lesson17_homework.Lesson17_homework_02 import TestHttpRequest
from openpyxl import load_workbook

class Do_Excel:
    def read_excel_data(self):
        wb = load_workbook('test_data.xlsx')  # 打开excel
        sheet = wb['Sheet1']  # 定位表单
        testdata = []  # 定义一个空列表        # --复盘！！！
        for i in range(1, sheet.max_row + 1):
            sub_data = {}  # 定义一个空字典         # --复盘！！！
            sub_data['url'] = sheet.cell(i, 1).value  # 第i行第1列数据
            sub_data['param'] = sheet.cell(i, 2).value  # 第i行第2列数据
            sub_data['http_method'] = sheet.cell(i, 3).value  # 第i行第3列数据
            sub_data['expected'] = sheet.cell(i, 4).value  # 第i行第4列数据
            testdata.append(sub_data)  # 把sub_data添加到列表中    # --复盘！！！
        return testdata
d = Do_Excel().read_excel_data()

suite = unittest.TestSuite()
for item in d:
    suite.addTest(TestHttpRequest(item['url'],
                                  eval(item['param']),
                                  item['http_method'],
                                  eval(item['expected']),
                                  'test_api'))
# # 将结果输出到txt文件中
# # with open('UnittestTextReport_v4.txt','a') as f:
# #     runner = unittest.TextTestRunner(stream = f,verbosity = 2)
#     runner.run(suite)


# 生成html报告
with open('test_report_v3.html','wb+') as f:
    runner = HTMLTestRunnerNew.HTMLTestRunner(stream = f,
                                           title='Http Request Report',
                                           description = 'Python12',
                                            tester = 'Hester',
                                           verbosity = 2)
    runner.run(suite)
