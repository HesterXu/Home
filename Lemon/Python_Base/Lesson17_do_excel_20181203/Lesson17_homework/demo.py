# -*- coding: utf-8 -*-
# @Time     : 2018/12/5/11:56
# @Author   : Hester Xu
# Email     : xuruizhu@yeah.net
# @File     : demo.py
# @Software : PyCharm


from Lemon.Python_Base.Lesson17_do_excel_20181203.Lesson17_homework.Lesson17_homework_01 import HttpRequest
from openpyxl import load_workbook

class Do_Excel:
    def read_data(self):
        wb = load_workbook('test_data.xlsx')     # 打开excel
        sheet = wb['Sheet1']                     # 定位表单
        global testdata
        testdata = []                            # 定义一个空列表        # --复盘！！！
        for i in range(1,sheet.max_row+1):
            sub_data = {}                        # 定义一个空字典         # --复盘！！！
            sub_data['url'] = sheet.cell(i, 1).value           # 第i行第1列数据
            sub_data['param'] = sheet.cell(i, 2).value      # 第i行第2列数据
            sub_data['http_method'] = sheet.cell(i, 3).value  # 第i行第3列数据
            sub_data['expected'] = sheet.cell(i, 4).value     # 第i行第4列数据
            testdata = testdata.append(sub_data)   # 把sub_data添加到列表中    # --复盘！！！
        return testdata











