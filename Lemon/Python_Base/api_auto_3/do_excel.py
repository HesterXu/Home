# -*- coding: utf-8 -*-
# @Time     : 2018/12/9/15:48
# @Author   : Hester Xu
# Email     : xuruizhu@yeah.net
# @File     : do_excel.py
# @Software : PyCharm

# 从excel表中拿数据

from openpyxl import load_workbook
from Lemon.Python_Base.api_auto_3.read_config import ReadConfig

class DoExcel:
    def get_data(self,file_name,sheet_name,button):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]

        test_data = []
        for i in range(2,sheet.max_row+1):
            row_data = {}
            row_data['case_id'] = sheet.cell(i,1).value
            row_data['title'] = sheet.cell(i, 2).value
            row_data['url'] = sheet.cell(i,3).value
            row_data['param'] = sheet.cell(i,4).value
            row_data['http_method'] = sheet.cell(i,5).value
            row_data['expected'] = sheet.cell(i,6).value
            test_data.append(row_data)

        final_data = []  # 最后的测试数据
        if button == 'all':
            final_data = test_data
        else:                          # 怎么读取button里的值，列表里的case_id？？？
            for item in test_data:
                if item['case_id'] in eval(button):
                    final_data.append(item)
        return final_data

if __name__ == '__main__':
    button = ReadConfig().read_config('case.configs', 'CASE', 'button')
    test_data = DoExcel().get_data('test_cases.xlsx','test_data',button)
    print(test_data)

